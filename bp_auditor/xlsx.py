#!/usr/bin/python3

import json

from pathlib import Path
from datetime import datetime
from contextlib import contextmanager

import openpyxl

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font

from .db import read_all_reports_from


@contextmanager
def open_calc_book(doc_location: str = 'report.xlsx'):
    # Create a new workbook
    wb = Workbook()

    yield wb

    # Save the Excel file

    p = Path(doc_location).resolve()
    if p.is_file():
        p.unlink()

    wb.save(doc_location)

COLOR_WHITE = 'eeeeee'

COLOR_HEADER_FG = COLOR_WHITE
COLOR_HEADER_BG = '333333'

COLOR_PASSED_FG = COLOR_WHITE
COLOR_PASSED_BG = '194d33'
COLORS_PASSED = [COLOR_PASSED_FG, COLOR_PASSED_BG]

COLOR_FAILED_FG = COLOR_WHITE
COLOR_FAILED_BG = 'f47373'
COLORS_FAILED = [COLOR_FAILED_FG, COLOR_FAILED_BG]

COLOR_CPU_GREEN = '33691e'
COLOR_CPU_YELLOW = 'ffa000'
COLOR_CPU_RED = 'f44336'

COLORS_CPU = [
    [COLOR_WHITE, COLOR_CPU_GREEN],
    [COLOR_WHITE, COLOR_CPU_YELLOW],
    [COLOR_WHITE, COLOR_CPU_RED],
]


def set_cell_color(ws, coord: str, fg_color: str, bg_color: str):
    # Create a new fill object with the gradient type set to 'path'
    ws[coord].fill = PatternFill(start_color=bg_color, fill_type='solid')
    ws[coord].font = Font(color=fg_color)


def get_sheet_or_create(wb, sheet_name: str):
    try:
        return wb.get_sheet_by_name(sheet_name)

    except KeyError:
        ws = wb.create_sheet(sheet_name)
        # Header
        ws['B1'] = 'URL'
        ws['C1'] = 'BP_JSON'
        ws['D1'] = 'TLS/SSL'
        ws['E1'] = 'P2P'
        ws['F1'] = 'History'
        ws['G1'] = 'CPU'

        for ord_col in range(ord('A'), ord('G') + 1):
            coord = chr(ord_col) + '1'
            set_cell_color(ws, coord, COLOR_HEADER_FG, COLOR_HEADER_BG)

        return ws



def produce_monthly_report(
    db_location: str,
    doc_location: str
):

    # Get the current month and year
    now = datetime.utcnow()
    current_month = now.month
    current_year = now.year

    # Get the timestamp for the first day of the current month at 00:00 UTC
    first_day_of_month = datetime(current_year, current_month, 1, 0, 0, 0)

    with open_calc_book(doc_location=doc_location) as workbook:
        # Get the default worksheet
        default_ws = workbook.active

        sheets = {}

        first_report_time = None
        last_report_time = None

        # Iterate through the rows and add data to cells
        for timestamp, reports in read_all_reports_from(db_location, first_day_of_month):
            if not first_report_time:
                first_report_time = datetime.fromisoformat(timestamp)

            last_report_time = datetime.fromisoformat(timestamp)

            for report in reports:
                san_url = report['url'].split('://')[1]
                san_url = san_url.split('/')[0]
                if san_url not in sheets:
                    sheets[san_url] = 2

                ws = get_sheet_or_create(workbook, san_url)

                ssl_report = ''
                ssl_passed = False
                if 'ssl_endpoints' in report:
                    for ssl_endpoint in report['ssl_endpoints']:
                        capabilities, url, tlsv = ssl_endpoint
                        ssl_passed = ssl_passed or (
                            ('query' in capabilities and tlsv in ['TLSv1.3', 'TLSv1.2']) or
                            ('full' in capabilities and tlsv in ['TLSv1.3', 'TLSv1.2'])
                        )
                        ssl_report += f'{json.dumps(capabilities)}, {tlsv}\n'

                p2p_report = ''
                p2p_passed = False
                if 'p2p_endpoints' in report:
                    for p2p_endpoint in report['p2p_endpoints']:
                        capabilities, url, status = p2p_endpoint
                        p2p_passed = p2p_passed or ('seed' in capabilities and status == 'ok')
                        p2p_report += f'{json.dumps(capabilities)}, {status}\n'

                hist_report = ''
                hist_passed = False
                if 'history' in report:
                    early = report['history']['early']
                    late = report['history']['late']

                    hist_report = 'early: '
                    early_passed = False
                    if len(early) == 2:
                        _id, block_num = early
                        hist_report += f'{block_num}\n'
                        early_passed = True
                    else:
                        hist_report += early + '\n'

                    hist_report += 'late: '
                    late_passed = False
                    if len(late) == 2:
                        _id, block_num = late
                        hist_report += f'{block_num}\n'
                        late_passed = True
                    else:
                        hist_report += late + '\n'

                    hist_passed = early_passed and late_passed

                if 'exception' in report:
                    ws.append([timestamp, report['exception']])
                else:
                    ws.append([
                        timestamp,
                        report['url'],
                        report['bp_json'],
                        ssl_report,
                        p2p_report,
                        hist_report,
                        report['cpu'] if 'cpu' in report else ''
                    ])

                    # Set alignment
                    for ord_col in range(ord('A'), ord('G') + 1):
                        col = chr(ord_col)
                        coord = col + str(sheets[san_url])

                        if col == 'D':
                            if ssl_passed:
                                set_cell_color(ws, coord, *COLORS_PASSED)
                            else:
                                set_cell_color(ws, coord, *COLORS_FAILED)

                        if col == 'E':
                            if p2p_passed:
                                set_cell_color(ws, coord, *COLORS_PASSED)
                            else:
                                set_cell_color(ws, coord, *COLORS_FAILED)

                        if col == 'F':
                            if hist_passed:
                                set_cell_color(ws, coord, *COLORS_PASSED)
                            else:
                                set_cell_color(ws, coord, *COLORS_FAILED)

                        if col == 'G':
                            cpu_val = ws[coord].value
                            if 'us' in cpu_val:
                                val = float(cpu_val.split(' ')[0])
                                rank = 2
                                if val >= 0 and val <= 300:
                                    rank = 0
                                elif val > 300 and val < 500:
                                    rank = 1
                                else:
                                    rank = 2

                                set_cell_color(ws, coord, *COLORS_CPU[rank])

                        ws[coord].alignment = Alignment(
                            horizontal='left', vertical='top')

                    sheets[san_url] += 1

        # Fix column size
        for sheet in sheets:
            ws = workbook[sheet]

            for col in range(ws.min_column, ws.max_column + 1):
                cell = ws[get_column_letter(col)+'2']
                length = max(20, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col)].width = length

            for row in range(ws.min_row, ws.max_row + 1):
                row = str(row)
                cell = ws['D'+row]
                if cell.value:
                    length = max(10, len(cell.value.split('\n')))
                    ws.row_dimensions[int(row)].height = length * 3

        ws = default_ws

        welcome_txt = 'Welcome to bp auto monitoring report v0.1a0\n'
        welcome_txt += f'generated at {str(now)}\nby "bp-auditor", a tool by\n\n - guille\n\n'
        welcome_txt += f'this report is for the days from \n\t\t{first_report_time} ___ to ___ {last_report_time}\n\n'
        diff = last_report_time - first_report_time
        diff = f'{diff.days} days, {diff.seconds//3600} hours, {(diff.seconds//60)%60} minutes, {diff.seconds%60} seconds'
        welcome_txt += f'a period of {diff}'

        ws.alignment = Alignment(horizontal='center', vertical='center')

        ws['A1'] = welcome_txt
        ws.column_dimensions['A'].width = 128
        ws.row_dimensions[1].height = 256

